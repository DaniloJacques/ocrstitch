
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Generator - Módulo para conversão de JSON para Excel
Integração com Power Automate para geração de planilhas
Author: Confrade Tech Solutions
Date: 2025
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """
    Classe para geração de planilhas Excel a partir dos resultados OCR
    """

    def __init__(self):
        """
        Inicializa o gerador de Excel
        """
        # Estilos para formatação
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.success_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.border = thin_border

        logger.info("Excel Generator inicializado")

    def load_json_results(self, json_file_path: str) -> Dict:
        """
        Carrega os resultados do arquivo JSON

        Args:
            json_file_path: Caminho para o arquivo JSON com resultados

        Returns:
            Dicionário com os dados carregados
        """
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            logger.info(f"JSON carregado: {len(data.get('resultados', []))} registros")
            return data

        except Exception as e:
            logger.error(f"Erro ao carregar JSON: {str(e)}")
            raise

    def create_summary_sheet(self, workbook: openpyxl.Workbook, data: Dict):
        """
        Cria a aba de resumo na planilha

        Args:
            workbook: Objeto workbook do openpyxl
            data: Dados do processamento OCR
        """
        # Remove a sheet padrão se existir
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        # Cria nova aba
        ws = workbook.create_sheet("📊 Resumo", 0)

        # Cabeçalho
        ws['A1'] = "RELATÓRIO DE PROCESSAMENTO OCR - LATAS"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:D1')

        # Informações gerais
        info_data = [
            ['Data/Hora do Processamento:', data.get('timestamp', 'N/A')],
            ['Total de Imagens Processadas:', data.get('total_processadas', 0)],
            ['Sucessos:', data.get('sucessos', 0)],
            ['Erros:', data.get('erros', 0)],
            ['Taxa de Sucesso:', f"{(data.get('sucessos', 0) / max(data.get('total_processadas', 1), 1) * 100):.1f}%"]
        ]

        start_row = 3
        for i, (label, value) in enumerate(info_data):
            ws[f'A{start_row + i}'] = label
            ws[f'B{start_row + i}'] = value
            ws[f'A{start_row + i}'].font = Font(bold=True)

        # Estatísticas por regras de negócio
        results = data.get('resultados', [])
        if results:
            ws[f'A{start_row + len(info_data) + 2}'] = "ESTATÍSTICAS POR OBSERVAÇÃO:"
            ws[f'A{start_row + len(info_data) + 2}'].font = Font(bold=True, size=12)

            # Conta observações
            obs_count = {}
            for result in results:
                obs = result.get('observacao', 'Sem observação')
                obs_count[obs] = obs_count.get(obs, 0) + 1

            stats_start = start_row + len(info_data) + 4
            for i, (obs, count) in enumerate(obs_count.items()):
                ws[f'A{stats_start + i}'] = obs
                ws[f'B{stats_start + i}'] = count

        # Formatação geral
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')

        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

        logger.info("Aba de resumo criada")

    def create_detailed_sheet(self, workbook: openpyxl.Workbook, data: Dict):
        """
        Cria a aba detalhada com todos os resultados

        Args:
            workbook: Objeto workbook do openpyxl
            data: Dados do processamento OCR
        """
        ws = workbook.create_sheet("📋 Detalhes")

        # Cabeçalhos das colunas
        headers = [
            'Arquivo',
            'Texto Extraído',
            'Texto Limpo',
            'Observação',
            'Status',
            'Data/Hora',
            'Caminho Completo'
        ]

        # Adiciona cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Adiciona dados
        results = data.get('resultados', [])
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=result.get('arquivo', ''))
            ws.cell(row=row, column=2, value=result.get('texto_extraido', ''))
            ws.cell(row=row, column=3, value=result.get('texto_limpo', ''))
            ws.cell(row=row, column=4, value=result.get('observacao', ''))
            ws.cell(row=row, column=5, value=result.get('status', ''))
            ws.cell(row=row, column=6, value=result.get('timestamp', ''))
            ws.cell(row=row, column=7, value=result.get('caminho_completo', ''))

            # Formatação condicional baseada no status
            status = result.get('status', '')
            fill = self.success_fill if status == 'sucesso' else self.error_fill

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')
                if status == 'erro':
                    cell.fill = self.error_fill

        # Ajusta largura das colunas
        column_widths = [20, 30, 30, 40, 10, 20, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Congela a primeira linha
        ws.freeze_panes = 'A2'

        logger.info(f"Aba detalhada criada com {len(results)} registros")

    def create_power_automate_sheet(self, workbook: openpyxl.Workbook, data: Dict):
        """
        Cria aba otimizada para Power Automate (formato simples)

        Args:
            workbook: Objeto workbook do openpyxl
            data: Dados do processamento OCR
        """
        ws = workbook.create_sheet("🤖 PowerAutomate")

        # Cabeçalhos simplificados para Power Automate
        headers = ['ID', 'Arquivo', 'TextoExtraido', 'Observacao', 'Status']

        # Adiciona cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        # Adiciona dados
        results = data.get('resultados', [])
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=row-1)  # ID sequencial
            ws.cell(row=row, column=2, value=result.get('arquivo', ''))
            ws.cell(row=row, column=3, value=result.get('texto_limpo', ''))
            ws.cell(row=row, column=4, value=result.get('observacao', ''))
            ws.cell(row=row, column=5, value=result.get('status', ''))

            # Aplica bordas
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.border

        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10

        logger.info("Aba para Power Automate criada")

    def generate_excel(self, json_file_path: str, excel_output_path: Optional[str] = None) -> str:
        """
        Gera planilha Excel completa a partir do JSON

        Args:
            json_file_path: Caminho para o arquivo JSON
            excel_output_path: Caminho de saída para Excel (opcional)

        Returns:
            Caminho do arquivo Excel gerado
        """
        try:
            # Carrega dados do JSON
            data = self.load_json_results(json_file_path)

            # Define caminho de saída se não fornecido
            if not excel_output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_output_path = f"relatorio_ocr_latas_{timestamp}.xlsx"

            # Cria workbook
            workbook = openpyxl.Workbook()

            # Cria todas as abas
            self.create_summary_sheet(workbook, data)
            self.create_detailed_sheet(workbook, data)
            self.create_power_automate_sheet(workbook, data)

            # Salva arquivo
            workbook.save(excel_output_path)
            logger.info(f"Planilha Excel salva em: {excel_output_path}")

            return excel_output_path

        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {str(e)}")
            raise

    def generate_csv_for_power_automate(self, json_file_path: str, csv_output_path: Optional[str] = None) -> str:
        """
        Gera arquivo CSV simples para integração com Power Automate

        Args:
            json_file_path: Caminho para o arquivo JSON
            csv_output_path: Caminho de saída para CSV (opcional)

        Returns:
            Caminho do arquivo CSV gerado
        """
        try:
            # Carrega dados
            data = self.load_json_results(json_file_path)

            # Define caminho de saída
            if not csv_output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                csv_output_path = f"dados_ocr_latas_{timestamp}.csv"

            # Prepara dados para DataFrame
            results = data.get('resultados', [])
            df_data = []

            for i, result in enumerate(results, 1):
                df_data.append({
                    'ID': i,
                    'Arquivo': result.get('arquivo', ''),
                    'TextoExtraido': result.get('texto_limpo', ''),
                    'Observacao': result.get('observacao', ''),
                    'Status': result.get('status', ''),
                    'Timestamp': result.get('timestamp', '')
                })

            # Cria DataFrame e salva CSV
            df = pd.DataFrame(df_data)
            df.to_csv(csv_output_path, index=False, encoding='utf-8-sig')

            logger.info(f"Arquivo CSV salvo em: {csv_output_path}")
            return csv_output_path

        except Exception as e:
            logger.error(f"Erro ao gerar CSV: {str(e)}")
            raise


def main():
    """
    Função principal para execução standalone
    """
    import argparse

    parser = argparse.ArgumentParser(description='Gerador de Excel/CSV para resultados OCR')
    parser.add_argument('json_file', help='Arquivo JSON com resultados OCR')
    parser.add_argument('-x', '--excel', help='Caminho para arquivo Excel de saída')
    parser.add_argument('-c', '--csv', help='Caminho para arquivo CSV de saída')
    parser.add_argument('--both', action='store_true', help='Gera tanto Excel quanto CSV')

    args = parser.parse_args()

    generator = ExcelGenerator()

    try:
        if args.excel or args.both:
            excel_path = generator.generate_excel(args.json_file, args.excel)
            print(f"✅ Excel gerado: {excel_path}")

        if args.csv or args.both:
            csv_path = generator.generate_csv_for_power_automate(args.json_file, args.csv)
            print(f"✅ CSV gerado: {csv_path}")

        if not args.excel and not args.csv and not args.both:
            # Padrão: gera Excel
            excel_path = generator.generate_excel(args.json_file)
            print(f"✅ Excel gerado: {excel_path}")

    except Exception as e:
        print(f"❌ Erro: {str(e)}")


if __name__ == "__main__":
    main()
